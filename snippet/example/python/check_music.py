#!/usr/bin/env python3
# -*- coding: utf8 -*-
#
# Support Python 3.6+
#
# Install:
# $ pip3 install six gevent requests sqlalchemy pymysql requests_html openpyxl
#
# Run:
# $ python check_music.py -h
#

import gevent.pool
import gevent.monkey
gevent.monkey.patch_all(Event=True, sys=True)
taskpool = gevent.pool.Pool(size=1000)
spawn = taskpool.spawn

import sys
import logging
import traceback

PY3, Unicode, Bytes = True, str, bytes
LOG = logging.getLogger()


def to_bytes(v, encoding="utf-8", **kwargs):
    if isinstance(v, Bytes):
        return v
    elif isinstance(v, Unicode):
        return v.encode(encoding)
    return to_bytes(str(v), encoding=encoding)


def to_unicode(v, encoding="utf-8", **kwargs):
    if isinstance(v, Bytes):
        return v.decode(encoding)
    elif isinstance(v, Unicode):
        return v
    return to_unicode(str(v), encoding=encoding)


to_str = to_unicode if PY3 else to_bytes
is_bytes = lambda s: isinstance(s, Bytes)
is_unicode = lambda s: isinstance(s, Unicode)
is_string = lambda s: isinstance(s, (Bytes, Unicode))


def init_logging(logger, level, log_file=None):
    fmt = "%(asctime)s - %(pathname)s - %(funcName)s - %(lineno)d - %(levelname)s - %(message)s"
    datefmt = "%Y-%m-%d %H:%M:%S"
    formatter = logging.Formatter(fmt=fmt, datefmt=datefmt)

    level = getattr(logging, level.upper())
    logger.setLevel(level)

    if log_file:
        from logging.handlers import TimedRotatingFileHandler
        handler = TimedRotatingFileHandler(log_file, when="midnight", interval=1, backupCount=30)
    else:
        handler = logging.StreamHandler()
    handler.setLevel(level)
    handler.setFormatter(formatter)

    logger.addHandler(handler)

##############################################################################
# Configuration
from argparse import ArgumentParser


# @Author: xgfone
# @Email: xgfone@126.com
class Configuration(object):
    class Group(object):
        def __init__(self, group_name):
            self.__name = group_name

        def __repr__(self):
            attrs = []
            for key, value in vars(self).items():
                if key != "_Group__name":
                    attrs.append("{0}={1}".format(key, value))
            return "{0}({1})".format(self.__class__.__name__, ", ".join(attrs))

        def __contains__(self, name):
            return hasattr(self, name)

        def __getattr__(self, name):
            e = "The group '{0}' has no the option '{1}'"
            raise AttributeError(e.format(self.__name, name))

        def __setitem__(self, name, value):
            setattr(self, name, value)

        def __getitem__(self, name):
            try:
                return getattr(self, name)
            except AttributeError:
                e = "The group '{0}' has no the option '{1}'"
                raise KeyError(e.format(self.__name, name))

        def items(self):
            d = vars(self)
            d.pop("_Group__name")
            return d.items()

    __slots__ = ["_default_group_name", "_default_group", "_allow_empty",
                 "_encoding", "_parsed", "_caches", "_opts", "_bool_true",
                 "_bool_false", "_py2", "_description", "_version"]

    def __init__(self, description=None, allow_empty=False, encoding="utf-8",
                 default_group="DEFAULT", version=None):
        """A simple configuration file parser based on the format INI.

        When an configuration option does not exist, for getting one default
        value, not raising an exception, please use the method of get(), or the
        builtin function of getattr().
        """

        self._parsed = False
        self._description = description
        self._default_group_name = default_group
        self._default_group = Configuration.Group(self._default_group_name)
        self._allow_empty = allow_empty
        self._encoding = encoding
        self._version = version if version else "Unknown"

        self._caches = {self._default_group_name: self._default_group}
        self._opts = {}

        self._bool_true = ["t", "1", "on", "true"]
        self._bool_false = ["f", "0", "off", "false"]

        try:
            "".decode()
        except AttributeError:
            self._py2 = False
        else:
            self._py2 = True

    def __getattr__(self, name):
        if not self._parsed:
            raise Exception("Not parsed")

        try:
            return self._caches[name]
        except KeyError:
            pass

        msg = "'{0}' object has no attribute '{1}'"
        raise AttributeError(msg.format(self.__class__.__name__, name))

    def __getitem__(self, name):
        if not self._parsed:
            raise Exception("Not parsed")

        _name = self._uniformize(name)
        try:
            return self._caches[_name]
        except KeyError:
            pass

        msg = "'{0}' has no key '{1}'"
        raise KeyError(msg.format(self.__class__.__name__, name))

    def __repr__(self):
        attrs = []
        for key, value in self._caches.items():
            attrs.append("{0}={1}".format(key, value))
        return "{0}({1})".format(self.__class__.__name__, ", ".join(attrs))

    def _set_group_opt(self, group_name, opt_name, opt_value, force=False):
        gname = group_name if group_name else self._default_group_name
        group = self._caches[gname]
        if hasattr(group, opt_name) and not force:
            e = "The group '{0}' has had the option of '{1}'"
            raise ValueError(e.format(gname, opt_name))
        setattr(self._caches[gname], opt_name, opt_value)

    def _register(self, name, parser, default=None, group=None, help=None, short=None):
        if self._parsed:
            raise Exception("Have been parsed")

        name = self._uniformize(name)
        group = self._uniformize(group if group else self._default_group_name)
        self._opts.setdefault(group, {})

        if name in self._opts[group]:
            raise KeyError("The option {0} has been regisetered".format(name))

        self._opts[group][name] = (parser, default, help, short)
        self._caches.setdefault(group, Configuration.Group(group))

    def _parse_int(self, value):
        return int(value)

    def _parse_float(self, value):
        return float(value)

    def _parse_bool(self, value):
        if isinstance(value, bool):
            return value
        elif not is_string(value):
            return bool(value)

        value = value.lower()
        if value in self._bool_true:
            return True
        elif value in self._bool_false:
            return False
        raise ValueError("invalid bool value '{0}'".format(value))

    def _parse_string(self, value):
        if self._py2:
            if isinstance(value, str):
                return value.decode(self._encoding)
        else:
            if not isinstance(value, str):
                return value.decode(self._encoding)
        return value

    def _parse_ints(self, value):
        return self._parse_list(self._parse_int, value)

    def _parse_strings(self, value):
        return self._parse_list(self._parse_string, value)

    def _parse_list(self, parser, value):
        if isinstance(value, (list, tuple)):
            vs = value
        else:
            vs = (v.strip() for v in value.split(",") if v.strip())
        return tuple((parser(v) for v in vs))

    def _uniformize(self, name):
        return name.replace("-", "_")

    def _unniformize(self, name):
        return name.replace("_", "-")

    def parsed(self):
        """Return True if it has been parsed, or False."""
        return self._parsed

    def parse_files(self, filenames=""):
        """Parse the INI configuration files.

        The argument is either a string standing for the path of the
        configuration file, or a list of them.
        """
        if self._parsed:
            raise Exception("Have been parsed")
        self._parsed = True

        if filenames:
            if not isinstance(filenames, (list, tuple)):
                filenames = self._parse_string(filenames).strip(", ").split(",")

            for filename in filenames:
                self._parse_file(filename)

        self._check_and_fix()

    def _check_and_fix(self):
        for gname, opts in self._opts.items():
            group = self._caches[gname]
            for name, opt in opts.items():
                if name in group:
                    continue
                elif opt[1] is not None or opt[0] == self._parse_bool:
                    self._set_group_opt(gname, name, opt[1])
                    continue

                if not self._allow_empty:
                    msg = "The option '{0}' in the group '{1}' has no value."
                    raise ValueError(msg.format(name, gname))

        # Set the options in the default group into self.
        group = self._caches.pop(self._default_group_name)
        for key, value in group.items():
            if key in self._caches:
                msg = "'{0}' had has the value '{1}'"
                raise ValueError(msg.format(self.__class__.__name__, key))
            self._caches[key] = value

    def _parse_file(self, filename):
        filename = str(filename)
        with open(filename) as f:
            lines = f.readlines()

        gname = self._default_group_name
        index, max_index = 0, len(lines)
        while index < max_index:
            line = self._parse_string(lines[index]).strip()
            index += 1

            # Comment
            if not line or line[0] in ("#", "=", ";"):
                continue

            # Group Section
            if line[0] == "[":
                if line[-1] != "]":
                    m = ("the format of the group is wrong, "
                         "which must start with [ and end with ]")
                    raise ValueError(m)
                _gname = line[1:-1]
                if not _gname:
                    raise ValueError("the group name is empty")
                if _gname not in self._caches:
                    continue
                gname = _gname
                continue

            # Group Option Values
            items = line.split("=", 1)
            if len(items) != 2:
                raise ValueError("the format is wrong, must contain '=': " + line)

            name, value = self._uniformize(items[0].strip()), items[1].strip()

            # Handle the continuation line
            if value[-1:] == "\\":
                values = [value.rstrip("\\").strip()]
                while index < max_index:
                    value = lines[index].strip()
                    values.append(value.rstrip("\\").strip())
                    index += 1
                    if value[-1:] != "\\":
                        break
                value = "\n".join(values)

            opt = self._opts[gname].get(name, None)
            if opt:
                self._set_group_opt(gname, name, opt[0](value))

    def register_bool(self, name, short=None, default=None, group=None, help=None):
        """Register the bool option.

        The value of this option will be parsed to the type of bool.
        """
        self._register(name, self._parse_bool, short=short, default=default,
                       group=group, help=help)

    def register_int(self, name, short=None, default=None, group=None, help=None):
        """Register the int option.

        The value of this option will be parsed to the type of int.
        """
        self._register(name, self._parse_int, short=short, default=default,
                       group=group, help=help)

    def register_float(self, name, short=None, default=None, group=None, help=None):
        """Register the float option.

        The value of this option will be parsed to the type of float.
        """
        self._register(name, self._parse_float, short=short, default=default,
                       group=group, help=help)

    def register_str(self, name, short=None, default=None, group=None, help=None):
        """Register the str option.

        The value of this option will be parsed to the type of str.
        """
        self._register(name, self._parse_string, short=short, default=default,
                       group=group, help=help)

    def register_int_list(self, name, short=None, default=None, group=None, help=None):
        """Register the int list option.

        The value of this option will be parsed to the type of int list.
        """
        self._register(name, self._parse_ints, short=short, default=default,
                       group=group, help=help)

    def register_str_list(self, name, short=None, default=None, group=None, help=None):
        """Register the string list option.

        The value of this option will be parsed to the type of string list.
        """
        self._register(name, self._parse_strings, short=short, default=default,
                       group=group, help=help)

    ###########################################################################
    # Parse CLI
    def parse(self, *args, **kwargs):
        return self.parse_cli(*args, **kwargs)

    def parse_cli(self, args=None, config_file_name="config-file"):
        """Parse the cli options."""
        if self._parsed:
            raise Exception("Have been parsed")
        self._parsed = True

        if args is None:
            args = sys.argv[1:]
        if not args:
            self._check_and_fix()
            return None

        gopts, args = self._parser_cli(args, description=self._description,
                                       config_file_name=config_file_name)
        if getattr(args, "version", False):
            print(self._version)
            sys.exit(0)

        if config_file_name:
            config_file = getattr(args, self._uniformize(config_file_name), "")
            for filename in config_file.split(","):
                filename = filename.strip()
                if filename:
                    self._parse_file(filename)

        for cli_opt, (gname, name) in gopts.items():
            opt = self._opts[gname][name]
            value = getattr(args, cli_opt, None)
            if value is not None:
                value = opt[0](value)
                if value != opt[1]:
                    self._set_group_opt(gname, name, value, force=True)

        self._check_and_fix()
        return args

    def _parser_cli(self, args, description=None, config_file_name=None):
        cli = ArgumentParser(description=description)
        if config_file_name:
            cli.add_argument("--" + config_file_name, default="",
                             help="The config file path.")
        cli.add_argument("--version", action="store_true",
                         help="Print the version and exit.")

        group_opts = {}
        for gname, opts in self._opts.items():
            if gname == self._default_group_name:
                group = cli
            else:
                group = cli.add_argument_group(gname)

            for name, (parser, default, help, short) in opts.items():
                action = None
                if parser == self._parse_bool:
                    action = "store_false" if default else "store_true"
                    default = False if default is None else default

                if gname == self._default_group_name:
                    opt_name = self._unniformize(name)
                    opt_key = self._uniformize(name)
                else:
                    opt_name = self._unniformize("{0}-{1}".format(gname, name))
                    opt_key = self._uniformize(opt_name)
                group_opts[opt_key] = (gname, name)
                short = "-" + short if short and short[0] != "-" else short
                names = [short, "--" + opt_name] if short else ["--" + opt_name]
                group.add_argument(*names, action=action, default=default, help=help)

        return group_opts, cli.parse_args(args=args)
# Configuration End
###############################################################################

###############################################################################
# Common
import requests
from urllib.parse import quote as qs_quote


def send_http_get(url, quote=True, use_key=False, co="?", timeout=5,
                  raise404=True, has_result=True, **ks):
    if ks:
        to = lambda v: qs_quote(to_str(v)) if quote else v
        ks = {k: to(v() if callable(v) else v) for k, v in ks.items() if v is not None}
        if use_key:
            url = co.join((url, "&".join(("%s=%s" % (k, v) for k, v in ks.items()))))
        else:
            url = url.format(**ks)
    resp = requests.get(url, timeout=timeout)
    status_code = resp.status_code
    if status_code == 404:
        if raise404:
            raise Exception("not found %s" % url)
        return None
    elif status_code == 200:
        if has_result:
            return resp.json()
        return None
    elif status_code == 204:
        return None
    raise OSError("%s: status_code=%s" % (url, status_code))
# Common End
###############################################################################

###############################################################################
# DB Common
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker, object_mapper
from sqlalchemy.sql.elements import TextClause

Iterator = object


class DB(object):
    """Manager the DB connection."""

    def __init__(self, write_connection, read_connection=None, autocommit=True,
                 expire_on_commit=False, echo=False, encoding=str("utf8"),
                 poolclass=None, pool=None, min_pool_size=2, max_pool_size=5,
                 pool_timeout=30, idle_timeout=3600):

        write_connection = self._fix_charset(write_connection, encoding)
        if read_connection:
            read_connection = self._fix_charset(read_connection, encoding)

        kwargs = {
            "echo": echo,
            "encoding": encoding,
            "poolclass": poolclass,
            "pool": pool,
            "pool_size": min_pool_size,
            "pool_timeout": pool_timeout if pool_timeout else None,
            "pool_recycle": idle_timeout,
            "max_overflow": max_pool_size - min_pool_size,
            "convert_unicode": True,
        }

        self._autocommit = autocommit
        self._expire_on_commit = expire_on_commit

        self._write_engine = self._create_engine(write_connection, kwargs)
        self._write_session_cls = self._get_session_cls(self._write_engine)

        if read_connection:
            self._read_engine = self._create_engine(read_connection, kwargs)
            self._read_session_cls = self._get_session_cls(self._read_engine)
        else:
            self._read_engine = self._write_engine
            self._read_session_cls = self._write_session_cls

    def _fix_charset(self, connection, encoding):
        if "mysql" in connection and "charset=" not in connection:
            if "?" in connection:
                return "%s&charset=%s" % (connection, encoding)
            return "%s?charset=%s" % (connection, encoding)
        return connection

    def _create_engine(self, connection, kwargs):
        if connection.startswith("sqlite:///"):
            kwargs.pop("pool_size", None)
            kwargs.pop("pool_timeout", None)
            kwargs.pop("max_overflow", None)
        return create_engine(connection, **kwargs)

    def _get_session_cls(self, engine):
        return sessionmaker(bind=engine, autocommit=self._autocommit,
                            expire_on_commit=self._expire_on_commit)

    def get_write_session(self):
        return self._write_session_cls()

    def get_read_session(self):
        return self._read_session_cls()

    def get_session(self):
        return self.get_write_session()

    def execute(self, sql, session=None, **kwargs):
        if not isinstance(sql, TextClause):
            sql = text(sql)
        return (session or self.get_session()).execute(sql, kwargs)

    def fetchall(self, sql, **kwargs):
        return self.execute(sql, self.get_read_session(), **kwargs).fetchall()

    def fetchone(self, sql, **kwargs):
        return self.execute(sql, self.get_read_session(), **kwargs).fetchone()

    def first(self, sql, **kwargs):
        return self.execute(sql, self.get_read_session(), **kwargs).first()


class ModelBase(Iterator):
    """Base class for models."""
    __tablename__ = ""
    __table_initialized__ = False

    def save(self, session):
        """Save this object."""

        # NOTE(boris-42): This part of code should be look like:
        #                       session.add(self)
        #                       session.flush()
        #                 But there is a bug in sqlalchemy and eventlet that
        #                 raises NoneType exception if there is no running
        #                 transaction and rollback is called. As long as
        #                 sqlalchemy has this bug we have to create transaction
        #                 explicitly.
        with session.begin(subtransactions=True):
            session.add(self)
            session.flush()

    def __repr__(self):
        attrs = ", ".join(("%s=%s" % (k, v) for k, v in self.items()))
        return "%s(%s)" % (self.__tablename__.title(), attrs)

    def __setitem__(self, key, value):
        setattr(self, key, value)

    def __getitem__(self, key):
        return getattr(self, key)

    def __contains__(self, key):
        # Don't use hasattr() because hasattr() catches any exception, not only
        # AttributeError. We want to passthrough SQLAlchemy exceptions
        # (ex: sqlalchemy.orm.exc.DetachedInstanceError).
        try:
            getattr(self, key)
        except AttributeError:
            return False
        else:
            return True

    def get(self, key, default=None):
        return getattr(self, key, default)

    def __iter__(self):
        columns = list(dict(object_mapper(self).columns).keys())
        return ModelIterator(self, iter(columns))

    def update(self, values):
        """Make the model object behave like a dict."""
        for k, v in values.items():
            setattr(self, k, v)

    def _as_dict(self):
        """Make the model object behave like a dict.
        Includes attributes from joins.
        """
        local = dict((key, value) for key, value in self)
        joined = dict([(k, v) for k, v in self.__dict__.items() if not k[0] == '_'])
        local.update(joined)
        return local

    def items(self):
        """Make the model object behave like a dict."""
        return self._as_dict().items()

    def keys(self):
        """Make the model object behave like a dict."""
        return [key for key, value in self.items()]


class ModelIterator(Iterator):
    def __init__(self, model, columns):
        self.model = model
        self.i = columns

    def __iter__(self):
        return self

    # In Python 3, __next__() has replaced next().
    def __next__(self):
        n = next(self.i)
        return n, getattr(self.model, n)

# DB Common End
###############################################################################

###############################################################################
# DB
from datetime import datetime
from sqlalchemy import Column, String, Boolean, Integer, DateTime
from sqlalchemy.sql import func, expression as expr
from sqlalchemy.ext.declarative import declarative_base

BASE = declarative_base()


class MusicBase(ModelBase):
    id = Column(Integer, primary_key=True)
    sid = Column(String(16), nullable=False)
    name = Column(String(128), nullable=False)
    singer = Column(String(128), nullable=False)
    ablum = Column(String(128), nullable=False)
    url = Column(String(128), nullable=False, server_default="")
    check = Column(Boolean, nullable=False, index=True, server_default=expr.text("0"))
    update_time = Column(DateTime, nullable=False, server_default=func.now())


class Music0(MusicBase, BASE):
    __tablename__ = "music0"


class Music1(MusicBase, BASE):
    __tablename__ = "music1"


class Music2(MusicBase, BASE):
    __tablename__ = "music2"


class Music3(MusicBase, BASE):
    __tablename__ = "music3"


class Music4(MusicBase, BASE):
    __tablename__ = "music4"


class Music5(MusicBase, BASE):
    __tablename__ = "music5"


class Music6(MusicBase, BASE):
    __tablename__ = "music6"


class Music7(MusicBase, BASE):
    __tablename__ = "music7"


class Music8(MusicBase, BASE):
    __tablename__ = "music8"


class Music9(MusicBase, BASE):
    __tablename__ = "music9"


class Music10(MusicBase, BASE):
    __tablename__ = "music10"


class Music11(MusicBase, BASE):
    __tablename__ = "music11"


class Music12(MusicBase, BASE):
    __tablename__ = "music12"


class Music13(MusicBase, BASE):
    __tablename__ = "music13"


class Music14(MusicBase, BASE):
    __tablename__ = "music14"


class Music15(MusicBase, BASE):
    __tablename__ = "music15"


class Music16(MusicBase, BASE):
    __tablename__ = "music16"


class Music17(MusicBase, BASE):
    __tablename__ = "music17"


class Music18(MusicBase, BASE):
    __tablename__ = "music18"


class Music19(MusicBase, BASE):
    __tablename__ = "music19"


class Music20(MusicBase, BASE):
    __tablename__ = "music20"

MusicModels = {
    "0": Music0,
    "1": Music1,
    "2": Music2,
    "3": Music3,
    "4": Music4,
    "5": Music5,
    "6": Music6,
    "7": Music7,
    "8": Music8,
    "9": Music9,
    "10": Music10,
    "11": Music11,
    "12": Music12,
    "13": Music13,
    "14": Music14,
    "15": Music15,
    "16": Music16,
    "17": Music17,
    "18": Music18,
    "19": Music19,
    "20": Music20,
}


class DBAPI(DB):
    def __init__(self, db_no, *args, **kwargs):
        super(DBAPI, self).__init__(*args, **kwargs)
        self._db_no = str(db_no)
        self._model = self._get_model(self._db_no)

    def create_tables(self):
        BASE.metadata.create_all(self._write_engine)

    def _get_model(self, sid):
        return MusicModels[sid[-1]]

    def get_unchecked_musics(self, num=1000):
        s = self.get_session()
        return s.query(self._model).filter_by(check=False).limit(num).all()

    def get_checked_musics(self):
        s = self.get_session()
        return s.query(self._model).filter_by(check=True).filter(self._model.url != "").all()

    def add_music(self, num, sid, name, singer, ablum, use_sid=False):
        m = self._get_model(sid[-1] if use_sid else str(num))
        m(sid=sid, name=name, singer=singer, ablum=ablum).save(self.get_session())

    def set_url(self, id, url):
        ks = {"update_time": datetime.now(), "check": True, "url": url}
        self.get_session().query(self._model).filter_by(id=id).update(ks)
# DB End
###############################################################################

import json
import os.path

from requests import ConnectionError, Timeout
from requests_html import HTMLSession

_SITES = {}


def register(name, *args, **kwargs):
    def decorate(cls):
        if name in _SITES:
            raise ValueError("The site '%s' has been registered" % name)
        _SITES[name] = cls(*args, **kwargs)
        return cls
    return decorate


def import_xlsx_into_db(dbapi, xlsx_path, columns=(0, 1, 2, 3), table_num=0):
    from openpyxl import load_workbook

    if not os.path.exists(xlsx_path):
        raise RuntimeError("The Excel file '%s' does not exist" % xlsx_path)

    def _import(i, num, id, name, ablum, singer):
        id = str(id.value)
        name = str(name.value)
        ablum = str(ablum.value)
        singer = str(singer.value)

        if id and name and ablum and singer:
            try:
                dbapi.add_music(num, id, name, singer, ablum, use_sid=not table_num)
            except Exception as err:
                LOG.error(err)

        LOG.debug("import row %s", i)

    ts = []
    num = 0
    id_i, name_i, ablum_i, singer_i = columns
    ws = load_workbook(xlsx_path).active
    for i, r in enumerate(ws.rows, 1):
        t = spawn(_import, i, num, r[id_i], r[name_i], r[ablum_i], r[singer_i])
        ts.append(t)
        num += 1
        if num >= table_num:
            num = 0
    for t in ts:
        t.join()


def export_to_excel(dbapi, xlsx_path):
    from openpyxl import Workbook

    if os.path.exists(xlsx_path):
        raise RuntimeError("The Excel file '%s' has existed" % xlsx_path)

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "SID"
    ws["B1"] = "Song Name"
    ws["C1"] = "Album Name"
    ws["D1"] = "Singer Name"
    ws["E1"] = "URL"
    for i, m in enumerate(dbapi.get_checked_musics(), 2):
        ws["A%s" % i] = m.sid
        ws["B%s" % i] = m.name
        ws["C%s" % i] = m.ablum
        ws["D%s" % i] = m.singer
        ws["E%s" % i] = m.url
    wb.save(xlsx_path)


def get_site(name):
    site = _SITES.get(name, None)
    if not site:
        raise ValueError("no support the site '%s'" % name)
    return site


def refactor_str(s):
    return ' '.join(s.lower().strip().split()).replace("'", "")


def refactor_ablum_name(s):
    return refactor_str(s.strip().lstrip("\u300a").rstrip("\u300b"))


def check_music(site_handler, dbapi, m, equal_test=False):
    try:
        infos = site_handler(m.name)
    except (ConnectionError, Timeout) as err:
        LOG.info("Timeout: sid=%s, name=%s, singer=%s, ablum=%s, err=%s",
                 m.sid, m.name, m.singer, m.ablum, err)
        return

    url = ""
    for info in infos:
        url = _check_music(info, m.name, m.singer, m.ablum, equal_test)
        if url:
            break

    try:
        dbapi.set_url(m.id, url)
    except Exception as err:
        LOG.error("Failed to set url: %s", err)

    if not infos:
        LOG.warning("NotFound: sid=%s, name=%s, singer=%s, ablum=%s",
                    m.sid, m.name, m.singer, m.ablum)
        return


def _check_music(info, song, singer, ablum, equal_test=False):
    song = refactor_str(song)
    singer = refactor_str(singer)
    ablum = refactor_ablum_name(ablum)

    if equal_test:
        if info["SongName"] != song:
            return ""
        if singer and info["SingerName"] != singer:
            return ""
        if ablum and info["AlbumName"] != ablum:
            return ""
    else:
        if singer and singer not in info["SingerName"]:
            return ""
        if ablum and ablum not in info["AlbumName"]:
            return ""

    return info["FileURL"]


def handle_musics(dbapi, site_handler, equal_test=False, num=1000, each=10):
    while True:
        musics = dbapi.get_unchecked_musics(num=num)
        if not musics:
            return
        while musics:
            ms = musics[:each]
            musics = musics[each:]
            tasks = []
            for m in ms:
                tasks.append(spawn(check_music, site_handler, dbapi, m, equal_test))
            for t in tasks:
                t.join()


@register("xiami")
class XiaMi:
    SEARCH_URL = "http://www.xiami.com/search?key="
    FILE_URL = ("http://www.xiami.com/play?ids=/song/playlist/id/{id}"
                "/object_name/default/object_id/0")

    def __init__(self):
        self._session = HTMLSession()

    def __call__(self, keyword):
        url = self.SEARCH_URL + qs_quote(keyword)
        html = self._session.get(url).html
        trs = html.find(".search_result_box .track_list tr")
        if not trs:
            return []

        results = []
        for tr in trs:
            tds = tr.find("td")
            if len(tds) < 5:
                continue

            a = tds[4].find("a", first=True)
            if not a:
                continue
            if a.attrs.get("title", "").strip() != "\u8bd5\u542c":
                continue

            onclick = a.attrs["onclick"].split("'")
            if len(onclick) < 2:
                continue
            id = onclick[1].strip()
            if not id:
                continue

            results.append({
                "FileURL": self.FILE_URL.format(id=id),
                "SongName": refactor_str(tds[1].text),
                "SingerName": refactor_str(tds[2].text),
                "AlbumName": refactor_ablum_name(tds[3].text),
            })
        return results


@register("kugou")
class KuGou:
    CALLBACK_KEY = "jQuery112406206328947895503"
    CALLBACK_LEN = len(CALLBACK_KEY) + 1
    FILE_URL = "http://www.kugou.com/song/#hash={hash}&album_id={album_id}"
    SEARCH_URL = ("http://songsearch.kugou.com/song_search_v2?callback={cb}&"
                  "page={page}&pagesize={size}&userid={userid}&keyword=")

    def __init__(self, page=1, page_size=30, userid=-1):
        self._url = self.SEARCH_URL.format(cb=self.CALLBACK_KEY, page=page,
                                           size=page_size, userid=userid)

    def _get_url(self, hash, album_id=""):
        return self.FILE_URL.format(hash=hash, album_id=album_id) if hash else ""

    def __call__(self, keyword):
        url = self._url + qs_quote(keyword)
        data = send_http_get(url, json=False)
        if not data:
            return None
        data = data[self.CALLBACK_LEN:-2]
        if not data:
            return None

        d = json.loads(data)
        if d["status"] != 1 or d["error_code"] != 0:
            msg = "The response error: status=%s, error_code=%s"
            raise ValueError(msg % (d["status"], d["error_code"]))

        results = []
        for m in d["data"]["lists"]:
            file1 = m.get("FileHash", None)
            file2 = m.get("HQFileHash", None)
            if not file1 and not file2:
                continue
            albumid = m.get("AlbumID", "")
            url = self._get_url(file1, albumid) or self._get_url(file2, albumid)
            m["FileURL"] = url
            m["AlbumName"] = refactor_ablum_name(m.get("AlbumName", ""))
            m["SongName"] = refactor_str(m["SongName"])
            m["SingerName"] = refactor_str(m["SingerName"])
            results.append(m)
        return results


def main(version="1.0.0"):
    conf = Configuration(description="Check the music.", version=version)
    conf.register_str("log_level", default="INFO",
                      help="The level of the log, such as debug, info, etc.")
    conf.register_str("log_file", default="", help="The file path of the log.")
    conf.register_bool("print_sql", help="Print the SQL statements.")
    conf.register_int("thread_num", default=0, help="The size of the coroutine pool.")
    conf.register_str("site", default="xiami", help="The site name, such as xiami.")
    conf.register_bool("equal_test", help="Check the music by the equal.")
    conf.register_int("db_no", default=0, help="The DB index number.")
    conf.register_int("db_pool_size", default=20, help="The max number of db conn pool.")
    conf.register_int("db_pool_timeout", default=30, help="The timeout of pool.")
    conf.register_str("db_conn", help="MySQL connection to query the music data.",
                      default="sqlite:///music.db")
    conf.register_bool("db_create", help="Initialize the Database.")
    conf.register_str("export_xlsx", default="", help="Export the result into a Excel file.")
    conf.register_str("import_xlsx", default="", help="Import the musics from Excel to DB.")
    conf.register_int("split_num", default=0, help="The number of the DB tables to split.")
    conf.register_int_list("xlsx_cols", default=(0, 1, 2, 3),
                           help="The index no of id, song name, ablum name, singer.")

    conf.parse()

    if conf.thread_num > 0:
        global taskpool, spawn
        taskpool = gevent.pool.Pool(size=conf.thread_num)
        spawn = taskpool.spawn
    timeout = None if conf.db_create else conf.db_pool_timeout

    init_logging(LOG, conf.log_level, conf.log_file)
    dbapi = DBAPI(conf.db_no, conf.db_conn, max_pool_size=conf.db_pool_size,
                  pool_timeout=timeout, idle_timeout=300, echo=conf.print_sql)

    if conf.db_create:
        dbapi.create_tables()
        return

    if conf.import_xlsx:
        import_xlsx_into_db(dbapi, conf.import_xlsx, conf.xlsx_cols, conf.split_num)
        return

    if conf.export_xlsx:
        export_to_excel(dbapi, conf.export_xlsx)
        return

    try:
        handle_musics(dbapi, get_site(conf.site), conf.equal_test)
    except Exception:
        LOG.error(traceback.format_exc())


if __name__ == "__main__":
    main()
